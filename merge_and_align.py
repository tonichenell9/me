#!/usr/bin/env python3
"""
Excel Merge and Align Utility
Merges data from two Excel spreadsheets and applies merge & centre formatting
to consecutive cells with identical values in specified columns.
"""

import logging
import re
from pathlib import Path
from typing import Optional, List, Union, Dict, Iterable, Tuple
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import pandas as pd

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

DEFAULT_OUTPUT_COLUMNS: List[str] = [
    # Note: user examples imply ISSUE NAME (A), SEDOL (B), ISIN (C)
    "ISSUE NAME",
    "SEDOL",
    "ISIN",
    "ISSUE COUNTRY NAME",
    "STRATEGY",
    "SUB STRATEGY",
    "FUND TYPE",
    "FUND NAME",
]


def _normalize_header(value: object) -> str:
    """
    Normalize a header into a comparable token:
    - stringified
    - lowercase
    - remove non-alphanumeric characters
    """
    if value is None:
        return ""
    s = str(value).strip().lower()
    # Treat underscores, spaces, dashes, punctuation as equivalent
    return re.sub(r"[^a-z0-9]+", "", s)


def _build_normalized_column_map(columns: Iterable[object]) -> Dict[str, str]:
    """
    Build a map from normalized header -> original column label (first occurrence).
    """
    mapping: Dict[str, str] = {}
    for col in columns:
        key = _normalize_header(col)
        if key and key not in mapping:
            mapping[key] = col  # keep original label as-is
    return mapping


def _extract_columns_case_insensitive(
    df: pd.DataFrame,
    canonical_columns: List[str],
    synonyms: Dict[str, List[str]],
) -> pd.DataFrame:
    """
    Extract a set of canonical columns from df by matching headers case-insensitively,
    ignoring punctuation/spacing differences.

    Any missing columns are created with empty values.
    """
    norm_map = _build_normalized_column_map(df.columns)

    out = pd.DataFrame(index=df.index)
    for canonical in canonical_columns:
        norm_targets = [_normalize_header(canonical)]
        for alt in synonyms.get(canonical, []):
            norm_targets.append(_normalize_header(alt))

        source_col = None
        for norm in norm_targets:
            if norm in norm_map:
                source_col = norm_map[norm]
                break

        if source_col is None:
            out[canonical] = pd.NA
        else:
            out[canonical] = df[source_col]

    return out


def consolidate_columns_and_group(
    file1_path: Union[str, Path],
    file2_path: Union[str, Path],
    output_path: Union[str, Path],
    output_columns: Optional[List[str]] = None,
    sheet1_name: Optional[str] = None,
    sheet2_name: Optional[str] = None,
    output_sheet_name: str = "MERGED",
) -> None:
    """
    Consolidate two spreadsheets by extracting a specific set of columns (case-insensitive),
    stacking rows, sorting by ISSUE NAME, and applying hierarchical merge/centre grouping.

    This matches the requested behavior:
    - Column titles output in UPPERCASE
    - Only the requested columns are included
    - Ordered left-to-right per `output_columns`
    - Sorted alphabetically by ISSUE NAME
    - Hierarchical vertical merges: for each column, merge consecutive duplicates
      within the grouping defined by all columns to its left.
    """
    file1_path = Path(file1_path)
    file2_path = Path(file2_path)
    output_path = Path(output_path)

    canonical = [c.upper() for c in (output_columns or DEFAULT_OUTPUT_COLUMNS)]

    # Minimal synonym list; normalization already handles casing/punctuation.
    # Add common variants where wording differs.
    synonyms: Dict[str, List[str]] = {
        "ISSUE NAME": ["Issue Name", "Issue name", "Security Name", "Instrument Name", "Name"],
        "ISSUE COUNTRY NAME": ["Issue Country", "Country", "Issue Country", "Issue Country Name"],
        "SUB STRATEGY": ["Sub-Strategy", "Sub Strategy", "Substrategy"],
        "FUND TYPE": ["Fund Type", "Fund type"],
        "FUND NAME": ["Fund Name", "Fund name"],
        "SEDOL": ["Sedol", "SEDOL CODE", "Sedol Code"],
        "ISIN": ["Isin", "ISIN CODE", "Isin Code"],
        "STRATEGY": ["Strategy"],
    }

    logger.info(f"Loading first file: {file1_path}")
    df1 = pd.read_excel(file1_path, sheet_name=sheet1_name or 0)
    logger.info(f"Loading second file: {file2_path}")
    df2 = pd.read_excel(file2_path, sheet_name=sheet2_name or 0)

    logger.info("Extracting canonical columns (case-insensitive match)")
    df1_out = _extract_columns_case_insensitive(df1, canonical, synonyms)
    df2_out = _extract_columns_case_insensitive(df2, canonical, synonyms)

    merged = pd.concat([df1_out, df2_out], ignore_index=True)

    # Drop rows that are completely empty across the selected columns
    merged = merged.dropna(how="all", subset=canonical)

    # Sort by ISSUE NAME (case-insensitive). If missing, leave unsorted.
    if "ISSUE NAME" in merged.columns:
        merged = merged.sort_values(
            by=["ISSUE NAME"] + [c for c in canonical if c != "ISSUE NAME"],
            key=lambda s: s.fillna("").astype(str).str.casefold(),
            kind="mergesort",  # stable
            na_position="last",
        )

    # Ensure final column order + uppercase headers
    merged = merged.reindex(columns=canonical)

    logger.info(f"Writing consolidated output to: {output_path}")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        merged.to_excel(writer, index=False, sheet_name=output_sheet_name)

    # Apply hierarchical merge/centre formatting using openpyxl
    wb = load_workbook(output_path)
    ws = wb[output_sheet_name]

    # Freeze header row for usability
    ws.freeze_panes = "A2"

    start_row = 2
    end_row = ws.max_row
    max_col = ws.max_column

    # Snapshot values before any merges (important: merging clears non-top cells)
    values: List[List[object]] = [
        [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        for r in range(start_row, end_row + 1)
    ]

    def is_blank(v: object) -> bool:
        return v is None or (isinstance(v, str) and v.strip() == "")

    def compute_hierarchical_merges(col_index_1based: int) -> List[Tuple[int, int]]:
        """
        Return row ranges (absolute sheet rows) to merge for the given column,
        within groups defined by all columns to its left.
        """
        merges: List[Tuple[int, int]] = []
        col0 = col_index_1based - 1  # 0-based into values row lists

        # Group by prefix key of columns to the left
        current_group_key = None
        group_start = 0  # index into values (0-based)

        for i in range(len(values) + 1):
            if i < len(values):
                group_key = tuple(values[i][0:col0])  # left columns
            else:
                group_key = None  # sentinel to flush last group

            if i == 0:
                current_group_key = group_key
                group_start = 0
                continue

            if group_key != current_group_key:
                # Process group [group_start, i-1] in terms of consecutive duplicates in col0
                run_value = values[group_start][col0] if group_start < len(values) else None
                run_start = group_start

                for j in range(group_start + 1, i + 1):
                    if j < i:
                        v = values[j][col0]
                    else:
                        v = None  # sentinel to flush run

                    if j == i or v != run_value:
                        run_end = j - 1
                        if run_end > run_start and not is_blank(run_value):
                            # Convert to absolute sheet row indices
                            merges.append((start_row + run_start, start_row + run_end))
                        run_value = v if j < i else None
                        run_start = j

                current_group_key = group_key
                group_start = i

        return merges

    logger.info("Applying hierarchical merge/centre formatting")
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        merge_ranges = compute_hierarchical_merges(col_idx)
        for r1, r2 in merge_ranges:
            merge_ref = f"{col_letter}{r1}:{col_letter}{r2}"
            try:
                ws.merge_cells(merge_ref)
                cell = ws.cell(row=r1, column=col_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            except Exception as e:
                logger.warning(f"Could not merge {merge_ref}: {e}")

    wb.save(output_path)
    logger.info("Consolidation + grouping completed successfully.")


def merge_excel_files(
    file1_path: Union[str, Path],
    file2_path: Union[str, Path],
    output_path: Union[str, Path],
    merge_type: str = 'vertical',
    sheet1_name: Optional[str] = None,
    sheet2_name: Optional[str] = None
) -> Workbook:
    """
    Merge data from two Excel files into a single workbook.
    
    Args:
        file1_path: Path to the first Excel file
        file2_path: Path to the second Excel file
        output_path: Path for the output merged file
        merge_type: 'vertical' (stack rows) or 'horizontal' (add columns side by side)
        sheet1_name: Optional name of sheet to read from file1 (defaults to active sheet)
        sheet2_name: Optional name of sheet to read from file2 (defaults to active sheet)
    
    Returns:
        The merged openpyxl Workbook object
    """
    file1_path = Path(file1_path)
    file2_path = Path(file2_path)
    output_path = Path(output_path)
    
    logger.info(f"Loading first file: {file1_path}")
    df1 = pd.read_excel(file1_path, sheet_name=sheet1_name or 0)
    
    logger.info(f"Loading second file: {file2_path}")
    df2 = pd.read_excel(file2_path, sheet_name=sheet2_name or 0)
    
    if merge_type == 'vertical':
        logger.info("Merging files vertically (stacking rows)")
        merged_df = pd.concat([df1, df2], ignore_index=True)
    elif merge_type == 'horizontal':
        logger.info("Merging files horizontally (side by side)")
        merged_df = pd.concat([df1, df2], axis=1)
    else:
        raise ValueError(f"Invalid merge_type: {merge_type}. Use 'vertical' or 'horizontal'")
    
    logger.info(f"Merged data shape: {merged_df.shape}")
    
    # Save to Excel using pandas first
    merged_df.to_excel(output_path, index=False)
    
    # Load with openpyxl for formatting
    workbook = load_workbook(output_path)
    logger.info(f"Merged file saved to: {output_path}")
    
    return workbook


def merge_and_centre_cells(
    workbook: Workbook,
    columns_to_merge: Optional[List[Union[int, str]]] = None,
    sheet_name: Optional[str] = None,
    start_row: int = 2,  # Default to 2 to skip header row
    end_row: Optional[int] = None
) -> Workbook:
    """
    Merge and centre consecutive cells with identical values in specified columns.
    
    Args:
        workbook: The openpyxl Workbook to process
        columns_to_merge: List of column indices (1-based) or letters to merge.
                         If None, all columns will be processed.
        sheet_name: Name of the sheet to process. Defaults to active sheet.
        start_row: Row to start merging from (1-based, default 2 to skip header)
        end_row: Row to end merging at. Defaults to last row with data.
    
    Returns:
        The modified Workbook object
    """
    if sheet_name:
        ws = workbook[sheet_name]
    else:
        ws = workbook.active
    
    if end_row is None:
        end_row = ws.max_row
    
    # Convert column references to indices if needed
    if columns_to_merge is None:
        columns_to_merge = list(range(1, ws.max_column + 1))
    else:
        processed_columns = []
        for col in columns_to_merge:
            if isinstance(col, str):
                # Convert column letter to index
                col_idx = openpyxl.utils.column_index_from_string(col)
                processed_columns.append(col_idx)
            else:
                processed_columns.append(col)
        columns_to_merge = processed_columns
    
    logger.info(f"Processing columns: {columns_to_merge}")
    logger.info(f"Row range: {start_row} to {end_row}")
    
    # Process each column
    for col_idx in columns_to_merge:
        col_letter = get_column_letter(col_idx)
        logger.info(f"Processing column {col_letter} (index {col_idx})")
        
        # Find consecutive cells with same values
        merge_ranges = []
        current_value = None
        range_start = start_row
        
        for row in range(start_row, end_row + 1):
            cell_value = ws.cell(row=row, column=col_idx).value
            
            if row == start_row:
                current_value = cell_value
                range_start = row
            elif cell_value != current_value:
                # End of current range
                if row - 1 > range_start:
                    merge_ranges.append((range_start, row - 1))
                current_value = cell_value
                range_start = row
        
        # Handle the last range
        if end_row > range_start and ws.cell(row=end_row, column=col_idx).value == current_value:
            merge_ranges.append((range_start, end_row))
        
        # Apply merge and centre to identified ranges
        for start, end in merge_ranges:
            merge_ref = f"{col_letter}{start}:{col_letter}{end}"
            logger.info(f"  Merging cells: {merge_ref}")
            
            try:
                ws.merge_cells(merge_ref)
                # Apply centre alignment to the merged cell
                merged_cell = ws.cell(row=start, column=col_idx)
                merged_cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True
                )
            except Exception as e:
                logger.warning(f"  Could not merge {merge_ref}: {e}")
    
    return workbook


def process_excel_merge_and_align(
    file1_path: Union[str, Path],
    file2_path: Union[str, Path],
    output_path: Union[str, Path],
    merge_type: str = 'vertical',
    columns_to_merge: Optional[List[Union[int, str]]] = None,
    skip_header: bool = True,
    sheet1_name: Optional[str] = None,
    sheet2_name: Optional[str] = None
) -> None:
    """
    Complete pipeline: merge two Excel files and apply merge & centre formatting.
    
    Args:
        file1_path: Path to the first Excel file
        file2_path: Path to the second Excel file  
        output_path: Path for the output file
        merge_type: 'vertical' (stack rows) or 'horizontal' (add columns)
        columns_to_merge: List of columns to apply merge & centre. 
                         Can be integers (1-based) or letters (e.g., 'A', 'B').
                         If None, all columns are processed.
        skip_header: If True, starts merging from row 2 (skips header row)
        sheet1_name: Name of sheet to read from file1
        sheet2_name: Name of sheet to read from file2
    """
    logger.info("=" * 60)
    logger.info("Starting Excel Merge and Align Process")
    logger.info("=" * 60)
    
    # Step 1: Merge the files
    workbook = merge_excel_files(
        file1_path=file1_path,
        file2_path=file2_path,
        output_path=output_path,
        merge_type=merge_type,
        sheet1_name=sheet1_name,
        sheet2_name=sheet2_name
    )
    
    # Step 2: Apply merge and centre formatting
    start_row = 2 if skip_header else 1
    workbook = merge_and_centre_cells(
        workbook=workbook,
        columns_to_merge=columns_to_merge,
        start_row=start_row
    )
    
    # Step 3: Save the final workbook
    output_path = Path(output_path)
    workbook.save(output_path)
    logger.info(f"Final output saved to: {output_path}")
    logger.info("Process completed successfully!")


def apply_merge_centre_to_existing(
    input_path: Union[str, Path],
    output_path: Optional[Union[str, Path]] = None,
    columns_to_merge: Optional[List[Union[int, str]]] = None,
    sheet_name: Optional[str] = None,
    skip_header: bool = True
) -> None:
    """
    Apply merge and centre formatting to an existing Excel file.
    
    Args:
        input_path: Path to the Excel file to process
        output_path: Path for the output file. If None, overwrites input file.
        columns_to_merge: List of columns to apply merge & centre.
                         Can be integers (1-based) or letters (e.g., 'A', 'B').
                         If None, all columns are processed.
        sheet_name: Name of sheet to process. Defaults to active sheet.
        skip_header: If True, starts merging from row 2 (skips header row)
    """
    input_path = Path(input_path)
    if output_path is None:
        output_path = input_path
    else:
        output_path = Path(output_path)
    
    logger.info(f"Loading workbook: {input_path}")
    workbook = load_workbook(input_path)
    
    start_row = 2 if skip_header else 1
    workbook = merge_and_centre_cells(
        workbook=workbook,
        columns_to_merge=columns_to_merge,
        sheet_name=sheet_name,
        start_row=start_row
    )
    
    workbook.save(output_path)
    logger.info(f"Output saved to: {output_path}")


# Example usage and CLI
if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Merge Excel files and apply merge & centre formatting to consecutive identical values"
    )
    
    subparsers = parser.add_subparsers(dest='command', help='Available commands')

    # Subcommand: consolidate - Extract, merge, sort, and hierarchical-group a fixed set of columns
    cons_parser = subparsers.add_parser(
        'consolidate',
        help='Extract same-named columns (case-insensitive), merge, sort, and group by merging duplicate blocks'
    )
    cons_parser.add_argument('file1', help='Path to first Excel file')
    cons_parser.add_argument('file2', help='Path to second Excel file')
    cons_parser.add_argument('output', help='Path for output file')
    cons_parser.add_argument(
        '--sheet1',
        help='Sheet name in file1 (default: first sheet)'
    )
    cons_parser.add_argument(
        '--sheet2',
        help='Sheet name in file2 (default: first sheet)'
    )
    cons_parser.add_argument(
        '--sheet-name',
        default='MERGED',
        help='Output sheet name (default: MERGED)'
    )
    cons_parser.add_argument(
        '--columns',
        nargs='*',
        help=(
            "Optional output columns (space-separated). "
            "If omitted, defaults to: "
            + ", ".join(DEFAULT_OUTPUT_COLUMNS)
        )
    )
    
    # Subcommand: merge - Merge two files and apply formatting
    merge_parser = subparsers.add_parser('merge', help='Merge two Excel files')
    merge_parser.add_argument('file1', help='Path to first Excel file')
    merge_parser.add_argument('file2', help='Path to second Excel file')
    merge_parser.add_argument('output', help='Path for output file')
    merge_parser.add_argument(
        '--merge-type', '-t',
        choices=['vertical', 'horizontal'],
        default='vertical',
        help='How to merge: vertical (stack rows) or horizontal (side by side)'
    )
    merge_parser.add_argument(
        '--columns', '-c',
        nargs='*',
        help='Columns to merge & centre (e.g., A B C or 1 2 3). If not specified, all columns.'
    )
    merge_parser.add_argument(
        '--no-skip-header',
        action='store_true',
        help='Include header row in merge operations'
    )
    
    # Subcommand: format - Apply merge & centre to existing file
    format_parser = subparsers.add_parser('format', help='Apply merge & centre to existing file')
    format_parser.add_argument('input', help='Path to Excel file')
    format_parser.add_argument('--output', '-o', help='Output path (default: overwrite input)')
    format_parser.add_argument(
        '--columns', '-c',
        nargs='*',
        help='Columns to merge & centre (e.g., A B C or 1 2 3). If not specified, all columns.'
    )
    format_parser.add_argument(
        '--sheet', '-s',
        help='Sheet name to process (default: active sheet)'
    )
    format_parser.add_argument(
        '--no-skip-header',
        action='store_true',
        help='Include header row in merge operations'
    )
    
    args = parser.parse_args()
    
    # Parse columns argument
    def parse_columns(cols):
        if cols is None:
            return None
        parsed = []
        for c in cols:
            try:
                parsed.append(int(c))
            except ValueError:
                parsed.append(c.upper())
        return parsed
    
    if args.command == 'merge':
        columns = parse_columns(args.columns)
        process_excel_merge_and_align(
            file1_path=args.file1,
            file2_path=args.file2,
            output_path=args.output,
            merge_type=args.merge_type,
            columns_to_merge=columns,
            skip_header=not args.no_skip_header
        )
    
    elif args.command == 'consolidate':
        consolidate_columns_and_group(
            file1_path=args.file1,
            file2_path=args.file2,
            output_path=args.output,
            output_columns=args.columns,
            sheet1_name=args.sheet1,
            sheet2_name=args.sheet2,
            output_sheet_name=args.sheet_name,
        )

    elif args.command == 'format':
        columns = parse_columns(args.columns)
        apply_merge_centre_to_existing(
            input_path=args.input,
            output_path=args.output,
            columns_to_merge=columns,
            sheet_name=args.sheet,
            skip_header=not args.no_skip_header
        )
    
    else:
        parser.print_help()
        print("\n" + "=" * 60)
        print("EXAMPLES:")
        print("=" * 60)
        print("\n0. Consolidate two files by matching columns case-insensitively, sorting, and grouping:")
        print("   python merge_and_align.py consolidate actions.xlsx accounts.xlsx output.xlsx")
        print("\n1. Merge two files vertically and apply merge & centre to all columns:")
        print("   python merge_and_align.py merge file1.xlsx file2.xlsx output.xlsx")
        print("\n2. Merge horizontally and only merge & centre columns A and B:")
        print("   python merge_and_align.py merge file1.xlsx file2.xlsx output.xlsx -t horizontal -c A B")
        print("\n3. Apply merge & centre to an existing file (columns 1, 2, 3):")
        print("   python merge_and_align.py format input.xlsx -c 1 2 3")
        print("\n4. Apply merge & centre to specific sheet:")
        print("   python merge_and_align.py format input.xlsx -o output.xlsx --sheet Sheet1 -c A B C")
