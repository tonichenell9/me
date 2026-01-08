#!/usr/bin/env python3
"""
Spreadsheet Consolidator (Simple Version)
Matches corporate actions data with accounts list using Entity ID,
then merges and applies hierarchical merge & center formatting.

HOW TO USE:
1. Edit the settings below
2. Run: python spreadsheet_consolidator_simple.py
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path

# =============================================================================
# ⬇️ EDIT THESE SETTINGS ⬇️
# =============================================================================

# Your input files (put them in the same folder as this script, or use full paths)
FILE_1 = "corporate_actions.xlsx"      # ← Corporate actions file (has Issue Name)
FILE_2 = "accounts_list.xlsx"          # ← Accounts list file (no Issue Name)

# Output file name
OUTPUT_FILE = "consolidated_output.xlsx"

# The column to match/join on (can be different names in each file)
LEFT_ON = "ENTITY ID"      # ← Column name in FILE_1 (corporate actions)
RIGHT_ON = "Entity ID"     # ← Column name in FILE_2 (accounts list)

# Merge type: "left", "right", "inner", "outer"
# "left" = keep all rows from accounts list that match corporate actions
HOW = "left"

# Columns to extract for the final output (in order from left to right)
# Note: ISSUE NAME comes from corporate actions file via the Entity ID match
COLUMNS_TO_EXTRACT = [
    "ISSUE NAME",
    "SEDOL",
    "ISIN",
    "ISSUE COUNTRY NAME",
    "STRATEGY",
    "SUB STRATEGY",
    "FUND TYPE",
    "FUND NAME",
]

# Set to False if you don't want cells merged
APPLY_MERGE_FORMATTING = True

# =============================================================================
# ⬆️ EDIT ABOVE THIS LINE ⬆️
# =============================================================================


def normalize_column_name(name):
    """Convert column name to uppercase for matching."""
    if name is None:
        return ''
    return str(name).strip().upper()


def find_matching_column(df, target_col):
    """Find column in dataframe (case-insensitive). Returns actual column name."""
    target_normalized = normalize_column_name(target_col)
    for col in df.columns:
        if normalize_column_name(col) == target_normalized:
            return col
    return None


def normalize_dataframe_columns(df):
    """Rename all columns to uppercase."""
    new_columns = {col: normalize_column_name(col) for col in df.columns}
    return df.rename(columns=new_columns)


def apply_hierarchical_merge(ws, start_row, end_row, columns):
    """Apply hierarchical merge and center-top formatting."""
    group_boundaries = [(start_row, end_row)]
    # Center horizontally, align to TOP vertically
    center_top_alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    
    for col_idx in columns:
        col_letter = get_column_letter(col_idx)
        merge_ranges = []
        new_boundaries = []
        
        for group_start, group_end in group_boundaries:
            if group_start > group_end:
                continue
            
            current_value = ws.cell(row=group_start, column=col_idx).value
            range_start = group_start
            
            for row in range(group_start, group_end + 1):
                cell_value = ws.cell(row=row, column=col_idx).value
                
                if cell_value != current_value:
                    if row - 1 >= range_start:
                        if row - 1 > range_start:
                            merge_ranges.append((range_start, row - 1))
                        new_boundaries.append((range_start, row - 1))
                    current_value = cell_value
                    range_start = row
            
            if range_start <= group_end:
                if group_end > range_start:
                    merge_ranges.append((range_start, group_end))
                new_boundaries.append((range_start, group_end))
        
        group_boundaries = new_boundaries
        
        for start, end in merge_ranges:
            try:
                ws.merge_cells(f"{col_letter}{start}:{col_letter}{end}")
                ws.cell(row=start, column=col_idx).alignment = center_top_alignment
            except:
                pass
        
        # Apply center-top alignment to all cells in the column
        for row in range(start_row, end_row + 1):
            try:
                ws.cell(row=row, column=col_idx).alignment = center_top_alignment
            except:
                pass


def apply_formatting(ws):
    """Apply header and cell formatting."""
    # Header formatting - no fill (clear), bold black text
    header_font = Font(bold=True, size=11, color="000000")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    ws.row_dimensions[1].height = 30
    
    # Data cell borders and center-top alignment
    center_top_alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = center_top_alignment
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 50)


def main():
    print("=" * 60)
    print("SPREADSHEET CONSOLIDATOR")
    print("=" * 60)
    
    # Check files exist
    if not Path(FILE_1).exists():
        print(f"\n❌ ERROR: Cannot find file '{FILE_1}'")
        print("   Make sure the file is in the same folder as this script,")
        print("   or update FILE_1 with the full path.")
        return
    
    if not Path(FILE_2).exists():
        print(f"\n❌ ERROR: Cannot find file '{FILE_2}'")
        print("   Make sure the file is in the same folder as this script,")
        print("   or update FILE_2 with the full path.")
        return
    
    # Load spreadsheets (keep original column names for matching)
    print(f"\n📂 Loading Corporate Actions: '{FILE_1}'...")
    df1 = pd.read_excel(FILE_1)
    print(f"   Found {len(df1)} rows, {len(df1.columns)} columns")
    print(f"   Columns: {list(df1.columns)}")
    
    print(f"\n📂 Loading Accounts List: '{FILE_2}'...")
    df2 = pd.read_excel(FILE_2)
    print(f"   Found {len(df2)} rows, {len(df2.columns)} columns")
    print(f"   Columns: {list(df2.columns)}")
    
    # Find the actual column names (case-insensitive search)
    left_col = find_matching_column(df1, LEFT_ON)
    right_col = find_matching_column(df2, RIGHT_ON)
    
    if left_col is None:
        print(f"\n❌ ERROR: Column '{LEFT_ON}' not found in corporate actions file")
        print(f"   Available columns: {list(df1.columns)}")
        return
    
    if right_col is None:
        print(f"\n❌ ERROR: Column '{RIGHT_ON}' not found in accounts list file")
        print(f"   Available columns: {list(df2.columns)}")
        return
    
    print(f"\n🔗 Matching: '{left_col}' (FILE_1) ↔ '{right_col}' (FILE_2)")
    print(f"   Merge type: {HOW}")
    
    # Get unique entity IDs from corporate actions
    entity_ids_in_corp_actions = df1[left_col].dropna().unique()
    print(f"   Found {len(entity_ids_in_corp_actions)} unique IDs in corporate actions")
    
    # Prepare corporate actions data for merge (just the columns we need)
    # Get ISSUE NAME column from corporate actions
    issue_name_col = find_matching_column(df1, "ISSUE NAME")
    
    if issue_name_col:
        # Create a lookup dataframe with Entity ID and Issue Name
        df1_lookup = df1[[left_col, issue_name_col]].drop_duplicates()
        df1_lookup.columns = ['_MERGE_KEY', 'ISSUE NAME']
        print(f"\n📋 Will add Issue Names from corporate actions")
    else:
        df1_lookup = df1[[left_col]].drop_duplicates()
        df1_lookup.columns = ['_MERGE_KEY']
        print(f"\n⚠️ No 'ISSUE NAME' column found in corporate actions")
    
    # Prepare accounts list for merge
    df2_for_merge = df2.copy()
    df2_for_merge['_MERGE_KEY'] = df2_for_merge[right_col]
    
    # Perform the merge (left join - keep accounts that match corporate actions)
    print(f"\n🔗 Performing {HOW} merge...")
    
    if HOW == "left":
        # Left join: keep all from df1_lookup (corporate actions IDs), match from df2
        # But we want accounts data, so we filter df2 to matching IDs
        merged_df = df2_for_merge[df2_for_merge['_MERGE_KEY'].isin(df1_lookup['_MERGE_KEY'])]
        merged_df = merged_df.merge(df1_lookup, on='_MERGE_KEY', how='left')
    else:
        merged_df = df2_for_merge.merge(df1_lookup, on='_MERGE_KEY', how=HOW)
    
    print(f"   Merged result: {len(merged_df)} rows")
    
    # Drop the merge key column
    merged_df = merged_df.drop(columns=['_MERGE_KEY'])
    
    # Normalize all column names to uppercase for extraction
    merged_df = normalize_dataframe_columns(merged_df)
    
    # Now extract the columns we want
    print(f"\n📊 Extracting columns...")
    columns_normalized = [normalize_column_name(c) for c in COLUMNS_TO_EXTRACT]
    
    # Check which columns are available
    available_columns = []
    for col in columns_normalized:
        if col in merged_df.columns:
            available_columns.append(col)
            print(f"   ✓ {col}")
        else:
            print(f"   ✗ {col} (not found)")
    
    if not available_columns:
        print("\n❌ ERROR: No columns found to extract!")
        return
    
    # Extract only the columns we want
    result_df = merged_df[available_columns].copy()
    
    # Remove completely empty rows
    result_df = result_df.dropna(how='all')
    print(f"\n   Final data: {len(result_df)} rows, {len(result_df.columns)} columns")
    
    # Sort by all columns (hierarchical sort)
    print(f"\n📊 Sorting data...")
    result_df = result_df.sort_values(by=available_columns, na_position='last')
    result_df = result_df.reset_index(drop=True)
    
    # Save to Excel
    print(f"\n💾 Saving to '{OUTPUT_FILE}'...")
    result_df.to_excel(OUTPUT_FILE, index=False, sheet_name='Consolidated Data')
    
    # Apply formatting
    print(f"\n🎨 Applying formatting...")
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    ws = wb.active
    
    apply_formatting(ws)
    
    if APPLY_MERGE_FORMATTING and ws.max_row > 1:
        print(f"   Merging duplicate cells...")
        apply_hierarchical_merge(ws, 2, ws.max_row, list(range(1, ws.max_column + 1)))
    
    wb.save(OUTPUT_FILE)
    
    print("\n" + "=" * 60)
    print(f"✅ DONE! Output saved to: {OUTPUT_FILE}")
    print("=" * 60)


if __name__ == "__main__":
    main()
