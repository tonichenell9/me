#!/usr/bin/env python3
"""
Spreadsheet Formatter
Takes an existing spreadsheet and applies:
1. Sorting (alphabetical by columns from left to right)
2. Hierarchical merge & center formatting for duplicate values
3. Cell formatting (borders, alignment)

HOW TO USE:
1. Edit the settings below
2. Run: python format_spreadsheet.py
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

# =============================================================================
# ⬇️ EDIT THESE SETTINGS ⬇️
# =============================================================================

# Input file (the merged spreadsheet to format)
INPUT_FILE = "merged_data.xlsx"

# Output file (leave same as input to overwrite, or use different name)
OUTPUT_FILE = "formatted_output.xlsx"

# Sheet name to process (use None for the first/active sheet)
SHEET_NAME = None

# Columns to sort by (in order). Use None to sort by all columns left to right.
# Example: ["ISSUE NAME", "SEDOL", "ISIN"] or None for all columns
SORT_BY_COLUMNS = None

# Apply hierarchical merge & center? Set to False to skip merging
APPLY_MERGE_FORMATTING = True

# Text alignment for data cells
HORIZONTAL_ALIGN = "center"   # Options: "left", "center", "right"
VERTICAL_ALIGN = "top"        # Options: "top", "center", "bottom"

# =============================================================================
# ⬆️ EDIT ABOVE THIS LINE ⬆️
# =============================================================================


def apply_hierarchical_merge(ws, start_row, end_row, columns, alignment):
    """Apply hierarchical merge and center formatting."""
    group_boundaries = [(start_row, end_row)]
    
    for col_idx in columns:
        col_letter = get_column_letter(col_idx)
        merge_ranges = []
        new_boundaries = []
        
        for group_start, group_end in group_boundaries:
            if group_start > group_end:
                continue
            
            current_value = ws.cell(row=group_start, column=col_idx).value
            range_start = group_start
            
            for row in range(group_start, group_end + 1):
                cell_value = ws.cell(row=row, column=col_idx).value
                
                if cell_value != current_value:
                    if row - 1 >= range_start:
                        if row - 1 > range_start:
                            merge_ranges.append((range_start, row - 1))
                        new_boundaries.append((range_start, row - 1))
                    current_value = cell_value
                    range_start = row
            
            if range_start <= group_end:
                if group_end > range_start:
                    merge_ranges.append((range_start, group_end))
                new_boundaries.append((range_start, group_end))
        
        group_boundaries = new_boundaries
        
        # Apply merges
        for start, end in merge_ranges:
            try:
                ws.merge_cells(f"{col_letter}{start}:{col_letter}{end}")
                ws.cell(row=start, column=col_idx).alignment = alignment
            except:
                pass
        
        # Apply alignment to all cells in the column
        for row in range(start_row, end_row + 1):
            try:
                ws.cell(row=row, column=col_idx).alignment = alignment
            except:
                pass


def apply_formatting(ws, data_alignment):
    """Apply header and cell formatting."""
    # Header formatting - no fill, bold black text
    header_font = Font(bold=True, size=11, color="000000")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Format header row
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    ws.row_dimensions[1].height = 30
    
    # Format data cells
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = data_alignment
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 50)


def main():
    print("=" * 60)
    print("SPREADSHEET FORMATTER")
    print("=" * 60)
    
    # Check file exists
    if not Path(INPUT_FILE).exists():
        print(f"\n❌ ERROR: Cannot find file '{INPUT_FILE}'")
        print("   Make sure the file is in the same folder as this script,")
        print("   or update INPUT_FILE with the full path.")
        return
    
    # Load spreadsheet
    print(f"\n📂 Loading '{INPUT_FILE}'...")
    df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME if SHEET_NAME else 0)
    print(f"   Found {len(df)} rows, {len(df.columns)} columns")
    print(f"   Columns: {list(df.columns)}")
    
    # Determine sort columns
    if SORT_BY_COLUMNS:
        sort_cols = [c for c in SORT_BY_COLUMNS if c in df.columns]
        if not sort_cols:
            print(f"\n⚠️ Warning: None of the specified sort columns found")
            sort_cols = list(df.columns)
    else:
        sort_cols = list(df.columns)
    
    # Sort data
    print(f"\n📊 Sorting by: {sort_cols}")
    df = df.sort_values(by=sort_cols, na_position='last')
    df = df.reset_index(drop=True)
    
    # Save to Excel
    print(f"\n💾 Saving to '{OUTPUT_FILE}'...")
    df.to_excel(OUTPUT_FILE, index=False, sheet_name='Formatted Data')
    
    # Load with openpyxl for formatting
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    ws = wb.active
    
    # Create alignment object
    data_alignment = Alignment(
        horizontal=HORIZONTAL_ALIGN,
        vertical=VERTICAL_ALIGN,
        wrap_text=True
    )
    
    # Apply basic formatting
    print(f"\n🎨 Applying formatting...")
    print(f"   Alignment: horizontal={HORIZONTAL_ALIGN}, vertical={VERTICAL_ALIGN}")
    apply_formatting(ws, data_alignment)
    
    # Apply hierarchical merge
    if APPLY_MERGE_FORMATTING and ws.max_row > 1:
        print(f"   Merging duplicate cells...")
        apply_hierarchical_merge(
            ws, 
            start_row=2, 
            end_row=ws.max_row, 
            columns=list(range(1, ws.max_column + 1)),
            alignment=data_alignment
        )
    
    # Save
    wb.save(OUTPUT_FILE)
    
    print("\n" + "=" * 60)
    print(f"✅ DONE! Output saved to: {OUTPUT_FILE}")
    print("=" * 60)


if __name__ == "__main__":
    main()
