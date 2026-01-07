#!/usr/bin/env python3
"""
Spreadsheet Consolidator
Merges data from two spreadsheets with different column name cases,
extracts specific columns, sorts data, and applies hierarchical 
merge & center formatting for duplicate values.
"""

import logging
from pathlib import Path
from typing import List, Optional, Union, Dict, Tuple
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# Default columns to extract (in order) - can be customized
DEFAULT_COLUMNS = [
    'ISSUE NAME',
    'SEDOL',
    'ISIN',
    'ISSUE COUNTRY NAME',
    'STRATEGY',
    'SUB STRATEGY',
    'FUND TYPE',
    'FUND NAME'
]


def normalize_column_name(name: str) -> str:
    """
    Normalize column name for case-insensitive matching.
    Converts to uppercase and strips whitespace.
    """
    if name is None:
        return ''
    return str(name).strip().upper()


def find_matching_column(df: pd.DataFrame, target_col: str) -> Optional[str]:
    """
    Find a column in the dataframe that matches the target column name
    (case-insensitive).
    
    Args:
        df: DataFrame to search
        target_col: Target column name to find
        
    Returns:
        The actual column name in the dataframe, or None if not found
    """
    target_normalized = normalize_column_name(target_col)
    
    for col in df.columns:
        if normalize_column_name(col) == target_normalized:
            return col
    
    return None


def extract_and_normalize_columns(
    df: pd.DataFrame,
    columns_to_extract: List[str]
) -> pd.DataFrame:
    """
    Extract specified columns from a dataframe using case-insensitive matching.
    Column names in the result are normalized to uppercase.
    
    Args:
        df: Source DataFrame
        columns_to_extract: List of column names to extract (case-insensitive)
        
    Returns:
        New DataFrame with extracted columns (uppercase names)
    """
    extracted_data = {}
    
    for target_col in columns_to_extract:
        actual_col = find_matching_column(df, target_col)
        if actual_col is not None:
            # Use uppercase version as the new column name
            normalized_name = normalize_column_name(target_col)
            extracted_data[normalized_name] = df[actual_col].values
            logger.info(f"  Found column '{actual_col}' -> '{normalized_name}'")
        else:
            logger.warning(f"  Column '{target_col}' not found in spreadsheet")
    
    return pd.DataFrame(extracted_data)


def merge_spreadsheets(
    file1_path: Union[str, Path],
    file2_path: Union[str, Path],
    columns_to_extract: List[str] = None,
    sheet1_name: Optional[Union[str, int]] = 0,
    sheet2_name: Optional[Union[str, int]] = 0
) -> pd.DataFrame:
    """
    Merge two spreadsheets by extracting matching columns (case-insensitive).
    
    Args:
        file1_path: Path to first spreadsheet (corporate actions data)
        file2_path: Path to second spreadsheet (accounts lists)
        columns_to_extract: List of column names to extract. If None, uses DEFAULT_COLUMNS
        sheet1_name: Sheet name or index for file1 (default: first sheet)
        sheet2_name: Sheet name or index for file2 (default: first sheet)
        
    Returns:
        Merged DataFrame with specified columns
    """
    if columns_to_extract is None:
        columns_to_extract = DEFAULT_COLUMNS.copy()
    
    file1_path = Path(file1_path)
    file2_path = Path(file2_path)
    
    logger.info(f"Loading spreadsheet 1: {file1_path}")
    df1 = pd.read_excel(file1_path, sheet_name=sheet1_name)
    logger.info(f"  Shape: {df1.shape}, Columns: {list(df1.columns)}")
    
    logger.info(f"Loading spreadsheet 2: {file2_path}")
    df2 = pd.read_excel(file2_path, sheet_name=sheet2_name)
    logger.info(f"  Shape: {df2.shape}, Columns: {list(df2.columns)}")
    
    # Extract and normalize columns from each spreadsheet
    logger.info("Extracting columns from spreadsheet 1...")
    extracted1 = extract_and_normalize_columns(df1, columns_to_extract)
    
    logger.info("Extracting columns from spreadsheet 2...")
    extracted2 = extract_and_normalize_columns(df2, columns_to_extract)
    
    # Ensure both dataframes have the same columns (in same order)
    # Add missing columns with NaN values
    all_columns = [normalize_column_name(c) for c in columns_to_extract]
    
    for col in all_columns:
        if col not in extracted1.columns:
            extracted1[col] = pd.NA
        if col not in extracted2.columns:
            extracted2[col] = pd.NA
    
    # Reorder columns
    extracted1 = extracted1[all_columns]
    extracted2 = extracted2[all_columns]
    
    # Merge (concatenate) the dataframes
    logger.info("Merging spreadsheets...")
    merged_df = pd.concat([extracted1, extracted2], ignore_index=True)
    
    # Remove completely empty rows
    merged_df = merged_df.dropna(how='all')
    
    logger.info(f"Merged data shape: {merged_df.shape}")
    
    return merged_df


def sort_dataframe(
    df: pd.DataFrame,
    sort_columns: List[str] = None
) -> pd.DataFrame:
    """
    Sort the dataframe by specified columns.
    
    Args:
        df: DataFrame to sort
        sort_columns: List of columns to sort by. If None, sorts by first column.
        
    Returns:
        Sorted DataFrame
    """
    if sort_columns is None:
        # Default: sort by first column (ISSUE NAME)
        sort_columns = [df.columns[0]] if len(df.columns) > 0 else []
    
    # Ensure sort columns exist
    valid_sort_cols = [c for c in sort_columns if c in df.columns]
    
    if valid_sort_cols:
        logger.info(f"Sorting by columns: {valid_sort_cols}")
        df = df.sort_values(by=valid_sort_cols, na_position='last')
        df = df.reset_index(drop=True)
    
    return df


def apply_hierarchical_merge_and_center(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    start_row: int = 2,
    end_row: Optional[int] = None,
    columns: Optional[List[int]] = None
) -> None:
    """
    Apply hierarchical merge and center formatting.
    
    For each column from left to right, merge consecutive cells with 
    identical values, but only within the merge boundaries established
    by columns to the left.
    
    This creates a hierarchical grouping effect where:
    - Column 1 (ISSUE NAME): Merged for all identical consecutive values
    - Column 2 (SEDOL): Merged only within each ISSUE NAME group
    - Column 3 (ISIN): Merged only within each SEDOL group
    - And so on...
    
    Args:
        ws: Worksheet to process
        start_row: First data row (after header)
        end_row: Last row to process. None = auto-detect
        columns: List of column indices (1-based). None = all columns
    """
    if end_row is None:
        end_row = ws.max_row
    
    if columns is None:
        columns = list(range(1, ws.max_column + 1))
    
    if end_row < start_row:
        logger.warning("No data rows to process")
        return
    
    # Track group boundaries for hierarchical merging
    # Each element is a tuple (group_start, group_end) defining the boundary
    group_boundaries = [(start_row, end_row)]
    
    center_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    
    for col_idx in columns:
        col_letter = get_column_letter(col_idx)
        logger.info(f"Processing column {col_letter} for merge & center")
        
        merge_ranges = []
        new_boundaries = []
        
        # Process each group boundary separately
        for group_start, group_end in group_boundaries:
            if group_start > group_end:
                continue
            
            # Find merge ranges within this group
            current_value = ws.cell(row=group_start, column=col_idx).value
            range_start = group_start
            
            for row in range(group_start, group_end + 1):
                cell_value = ws.cell(row=row, column=col_idx).value
                
                if cell_value != current_value:
                    # End current range, record new boundary
                    if row - 1 >= range_start:
                        if row - 1 > range_start:
                            merge_ranges.append((range_start, row - 1))
                        new_boundaries.append((range_start, row - 1))
                    current_value = cell_value
                    range_start = row
            
            # Handle the last range in this group
            if range_start <= group_end:
                if group_end > range_start:
                    merge_ranges.append((range_start, group_end))
                new_boundaries.append((range_start, group_end))
        
        # Update boundaries for next column
        group_boundaries = new_boundaries
        
        # Apply merges
        merge_count = 0
        for start, end in merge_ranges:
            merge_ref = f"{col_letter}{start}:{col_letter}{end}"
            
            try:
                ws.merge_cells(merge_ref)
                merged_cell = ws.cell(row=start, column=col_idx)
                merged_cell.alignment = center_alignment
                merge_count += 1
                logger.debug(f"  Merged: {merge_ref}")
            except Exception as e:
                logger.warning(f"  Could not merge {merge_ref}: {e}")
        
        if merge_count > 0:
            logger.info(f"  Applied {merge_count} merge(s) in column {col_letter}")
        
        # Apply center alignment to all cells in the column
        for row in range(start_row, end_row + 1):
            try:
                cell = ws.cell(row=row, column=col_idx)
                cell.alignment = center_alignment
            except:
                pass  # Cell might be part of a merge


def apply_header_formatting(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """
    Apply formatting to the header row.
    """
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Set row height
    ws.row_dimensions[1].height = 30


def apply_data_formatting(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    start_row: int = 2
) -> None:
    """
    Apply basic formatting to data cells.
    """
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(start_row, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border


def auto_adjust_column_widths(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """
    Auto-adjust column widths based on content.
    """
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # Set width with some padding, min 10, max 50
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def consolidate_spreadsheets(
    file1_path: Union[str, Path],
    file2_path: Union[str, Path],
    output_path: Union[str, Path],
    columns_to_extract: List[str] = None,
    sort_by: List[str] = None,
    apply_merge_format: bool = True,
    sheet1_name: Optional[Union[str, int]] = 0,
    sheet2_name: Optional[Union[str, int]] = 0
) -> str:
    """
    Main function to consolidate two spreadsheets.
    
    Args:
        file1_path: Path to first spreadsheet (e.g., corporate actions data)
        file2_path: Path to second spreadsheet (e.g., accounts lists)
        output_path: Path for the output consolidated spreadsheet
        columns_to_extract: List of column names to extract (case-insensitive).
                           Columns will be in this order in the output.
                           If None, uses DEFAULT_COLUMNS.
        sort_by: List of columns to sort by. If None, sorts by first column.
        apply_merge_format: If True, applies hierarchical merge & center formatting
        sheet1_name: Sheet name or index for file1
        sheet2_name: Sheet name or index for file2
        
    Returns:
        Path to the output file
    """
    logger.info("=" * 70)
    logger.info("SPREADSHEET CONSOLIDATION STARTED")
    logger.info("=" * 70)
    
    if columns_to_extract is None:
        columns_to_extract = DEFAULT_COLUMNS.copy()
    
    output_path = Path(output_path)
    
    # Step 1: Merge spreadsheets
    logger.info("\nStep 1: Merging spreadsheets...")
    merged_df = merge_spreadsheets(
        file1_path=file1_path,
        file2_path=file2_path,
        columns_to_extract=columns_to_extract,
        sheet1_name=sheet1_name,
        sheet2_name=sheet2_name
    )
    
    # Step 2: Sort data
    logger.info("\nStep 2: Sorting data...")
    if sort_by is None:
        # Default sort by all columns in order (hierarchical sort)
        sort_by = [normalize_column_name(c) for c in columns_to_extract]
    merged_df = sort_dataframe(merged_df, sort_by)
    
    # Step 3: Write to Excel
    logger.info("\nStep 3: Writing to Excel...")
    merged_df.to_excel(output_path, index=False, sheet_name='Consolidated Data')
    
    # Step 4: Apply formatting
    logger.info("\nStep 4: Applying formatting...")
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    
    # Apply header formatting
    apply_header_formatting(ws)
    
    # Apply data formatting (borders)
    apply_data_formatting(ws)
    
    # Apply hierarchical merge and center
    if apply_merge_format and ws.max_row > 1:
        logger.info("\nStep 5: Applying hierarchical merge & center formatting...")
        apply_hierarchical_merge_and_center(
            ws=ws,
            start_row=2,
            end_row=ws.max_row,
            columns=list(range(1, ws.max_column + 1))
        )
    
    # Auto-adjust column widths
    auto_adjust_column_widths(ws)
    
    # Save workbook
    wb.save(output_path)
    logger.info(f"\nOutput saved to: {output_path}")
    logger.info("=" * 70)
    logger.info("CONSOLIDATION COMPLETE!")
    logger.info("=" * 70)
    
    return str(output_path)


def create_sample_spreadsheets(output_dir: Union[str, Path] = ".") -> Tuple[str, str]:
    """
    Create sample spreadsheets for testing.
    
    Args:
        output_dir: Directory to save sample files
        
    Returns:
        Tuple of (file1_path, file2_path)
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Sample data for spreadsheet 1 (UPPERCASE columns - Corporate Actions)
    data1 = {
        'ISSUE NAME': ['Apple Inc', 'Apple Inc', 'Apple Inc', 'Microsoft Corp', 'Microsoft Corp', 
                       'Google LLC', 'Google LLC', 'Amazon Inc', 'Tesla Inc', 'Tesla Inc'],
        'SEDOL': ['B0YQ5W0', 'B0YQ5W0', 'B0YQ5W1', '2588173', '2588173',
                  'B7TL820', 'B7TL821', 'B0R64J4', 'B616C79', 'B616C80'],
        'ISIN': ['US0378331005', 'US0378331005', 'US0378331006', 'US5949181045', 'US5949181045',
                 'US02079K3059', 'US02079K3060', 'US0231351067', 'US88160R1014', 'US88160R1015'],
        'ISSUE COUNTRY NAME': ['United States', 'United States', 'United States', 'United States', 'United States',
                               'United States', 'United States', 'United States', 'United States', 'United States'],
        'STRATEGY': ['Growth', 'Growth', 'Growth', 'Value', 'Value',
                     'Growth', 'Growth', 'Growth', 'Growth', 'Growth'],
        'SUB STRATEGY': ['Tech Large Cap', 'Tech Large Cap', 'Tech Large Cap', 'Tech Large Cap', 'Tech Large Cap',
                         'Tech Large Cap', 'Tech Large Cap', 'Consumer', 'EV', 'EV'],
        'FUND TYPE': ['Equity', 'Equity', 'Equity', 'Equity', 'Equity',
                      'Equity', 'Equity', 'Equity', 'Equity', 'Equity'],
        'FUND NAME': ['Tech Growth Fund', 'Tech Growth Fund', 'Innovation Fund', 'Value Fund', 'Value Fund',
                      'Tech Growth Fund', 'Innovation Fund', 'Consumer Fund', 'EV Fund', 'EV Fund'],
        'EXTRA_COL_1': ['X1', 'X2', 'X3', 'X4', 'X5', 'X6', 'X7', 'X8', 'X9', 'X10']  # Should be excluded
    }
    
    # Sample data for spreadsheet 2 (Sentence case columns - Accounts List)
    data2 = {
        'Issue Name': ['Apple Inc', 'Apple Inc', 'Netflix Inc', 'Netflix Inc', 'Meta Platforms',
                       'Meta Platforms', 'Nvidia Corp', 'Nvidia Corp', 'Nvidia Corp', 'AMD Inc'],
        'Sedol': ['B0YQ5W0', 'B0YQ5W2', 'B07GKM0', 'B07GKM1', 'B7TL4S0',
                  'B7TL4S1', 'B0CYTT0', 'B0CYTT0', 'B0CYTT1', 'B0P2FQ2'],
        'Isin': ['US0378331005', 'US0378331007', 'US64110L1061', 'US64110L1062', 'US30303M1027',
                 'US30303M1028', 'US67066G1040', 'US67066G1040', 'US67066G1041', 'US0079031078'],
        'Issue Country Name': ['United States', 'United States', 'United States', 'United States', 'United States',
                               'United States', 'United States', 'United States', 'United States', 'United States'],
        'Strategy': ['Growth', 'Growth', 'Growth', 'Growth', 'Growth',
                     'Growth', 'Growth', 'Growth', 'Growth', 'Value'],
        'Sub Strategy': ['Tech Large Cap', 'Tech Large Cap', 'Entertainment', 'Entertainment', 'Social Media',
                         'Social Media', 'Semiconductors', 'Semiconductors', 'Semiconductors', 'Semiconductors'],
        'Fund Type': ['Equity', 'Equity', 'Equity', 'Equity', 'Equity',
                      'Equity', 'Equity', 'Equity', 'Equity', 'Equity'],
        'Fund Name': ['Tech Growth Fund', 'Innovation Fund', 'Entertainment Fund', 'Entertainment Fund', 'Social Fund',
                      'Social Fund', 'Chip Fund', 'Chip Fund', 'Innovation Fund', 'Value Fund'],
        'Account Number': ['A001', 'A002', 'A003', 'A004', 'A005', 'A006', 'A007', 'A008', 'A009', 'A010']  # Should be excluded
    }
    
    df1 = pd.DataFrame(data1)
    df2 = pd.DataFrame(data2)
    
    file1_path = output_dir / 'corporate_actions_sample.xlsx'
    file2_path = output_dir / 'accounts_list_sample.xlsx'
    
    df1.to_excel(file1_path, index=False, sheet_name='Corporate Actions')
    df2.to_excel(file2_path, index=False, sheet_name='Accounts')
    
    logger.info(f"Sample spreadsheet 1 created: {file1_path}")
    logger.info(f"Sample spreadsheet 2 created: {file2_path}")
    
    return str(file1_path), str(file2_path)


# Command-line interface
if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Consolidate two spreadsheets with different column name cases",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Consolidate two spreadsheets with default columns
  python spreadsheet_consolidator.py file1.xlsx file2.xlsx output.xlsx

  # Specify custom columns to extract
  python spreadsheet_consolidator.py file1.xlsx file2.xlsx output.xlsx \\
      --columns "ISSUE NAME" "SEDOL" "ISIN" "STRATEGY"

  # Create sample spreadsheets for testing
  python spreadsheet_consolidator.py --create-samples

  # Process without merge formatting
  python spreadsheet_consolidator.py file1.xlsx file2.xlsx output.xlsx --no-merge
        """
    )
    
    parser.add_argument('file1', nargs='?', help='Path to first spreadsheet')
    parser.add_argument('file2', nargs='?', help='Path to second spreadsheet')
    parser.add_argument('output', nargs='?', help='Path for output file')
    
    parser.add_argument(
        '--columns', '-c',
        nargs='+',
        help='Column names to extract (in order). Case-insensitive.'
    )
    
    parser.add_argument(
        '--sort-by', '-s',
        nargs='+',
        help='Column names to sort by (default: all columns in order)'
    )
    
    parser.add_argument(
        '--no-merge',
        action='store_true',
        help='Skip merge & center formatting'
    )
    
    parser.add_argument(
        '--sheet1',
        default=0,
        help='Sheet name or index for file1 (default: first sheet)'
    )
    
    parser.add_argument(
        '--sheet2',
        default=0,
        help='Sheet name or index for file2 (default: first sheet)'
    )
    
    parser.add_argument(
        '--create-samples',
        action='store_true',
        help='Create sample spreadsheets for testing'
    )
    
    parser.add_argument(
        '--sample-dir',
        default='.',
        help='Directory for sample files (default: current directory)'
    )
    
    args = parser.parse_args()
    
    if args.create_samples:
        file1, file2 = create_sample_spreadsheets(args.sample_dir)
        print(f"\nSample files created:")
        print(f"  1. {file1}")
        print(f"  2. {file2}")
        print(f"\nTo consolidate them, run:")
        print(f"  python spreadsheet_consolidator.py \"{file1}\" \"{file2}\" consolidated_output.xlsx")
    elif args.file1 and args.file2 and args.output:
        # Parse sheet arguments (convert to int if numeric)
        sheet1 = args.sheet1 if isinstance(args.sheet1, int) else (int(args.sheet1) if str(args.sheet1).isdigit() else args.sheet1)
        sheet2 = args.sheet2 if isinstance(args.sheet2, int) else (int(args.sheet2) if str(args.sheet2).isdigit() else args.sheet2)
        
        consolidate_spreadsheets(
            file1_path=args.file1,
            file2_path=args.file2,
            output_path=args.output,
            columns_to_extract=args.columns,
            sort_by=args.sort_by,
            apply_merge_format=not args.no_merge,
            sheet1_name=sheet1,
            sheet2_name=sheet2
        )
    else:
        parser.print_help()
        print("\n" + "=" * 70)
        print("DEFAULT COLUMNS (in order):")
        print("=" * 70)
        for i, col in enumerate(DEFAULT_COLUMNS, 1):
            print(f"  {i}. {col}")
        print("\nNote: Column matching is case-insensitive.")
        print("Output columns will be in UPPERCASE.")
