#!/usr/bin/env python3
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
import argparse
import sys
import os

def merge_and_format(file1_path, file2_path, output_path, merge_col_name=None, merge_col_idx=None):
    """
    Merges two Excel files and merges/centers cells in a specific column 
    where values are identical and adjacent.
    """
    
    print(f"Reading {file1_path}...")
    try:
        df1 = pd.read_excel(file1_path)
    except Exception as e:
        print(f"Error reading {file1_path}: {e}")
        return False

    print(f"Reading {file2_path}...")
    try:
        df2 = pd.read_excel(file2_path)
    except Exception as e:
        print(f"Error reading {file2_path}: {e}")
        return False

    # Concatenate the dataframes
    print("Merging data...")
    merged_df = pd.concat([df1, df2], ignore_index=True)

    # Save to Excel first using pandas to get the basic structure
    print(f"Saving intermediate result to {output_path}...")
    try:
        merged_df.to_excel(output_path, index=False)
    except Exception as e:
        print(f"Error saving to {output_path}: {e}")
        return False

    # Now open with openpyxl to apply formatting
    print("Applying Merge and Center formatting...")
    try:
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        
        # Determine the column index (1-based for openpyxl)
        col_index = None
        
        if merge_col_name:
            # Find column by name
            for idx, cell in enumerate(ws[1], 1):
                if cell.value == merge_col_name:
                    col_index = idx
                    break
            if col_index is None:
                print(f"Column '{merge_col_name}' not found in the merged file.")
                return False
        elif merge_col_idx is not None:
            col_index = int(merge_col_idx)
            if col_index < 1 or col_index > ws.max_column:
                print(f"Column index {col_index} is out of bounds.")
                return False
        else:
            print("No merge column specified. Skipping Merge and Center.")
            return True

        # Iterate through rows to find identical adjacent values
        # ws.iter_rows starts at row 1 (header), so data starts at row 2
        
        start_row = 2
        max_row = ws.max_row
        
        if max_row < 2:
            print("Not enough data to merge.")
            return True

        current_val = ws.cell(row=start_row, column=col_index).value
        merge_start = start_row

        # We iterate from the second data row up to max_row + 1 to handle the last group
        for row in range(start_row + 1, max_row + 2):
            if row <= max_row:
                cell_val = ws.cell(row=row, column=col_index).value
            else:
                cell_val = object() # Sentinel for end of loop

            if cell_val != current_val:
                # Value changed, check if we need to merge previous block
                if row - merge_start > 1:
                    # Merge cells
                    ws.merge_cells(start_row=merge_start, start_column=col_index, 
                                   end_row=row-1, end_column=col_index)
                    
                    # Center the merged cell (top-left cell of the merge)
                    top_left_cell = ws.cell(row=merge_start, column=col_index)
                    top_left_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Reset for next block
                current_val = cell_val
                merge_start = row

        wb.save(output_path)
        print(f"Successfully saved merged and formatted file to {output_path}")
        return True

    except Exception as e:
        print(f"Error during formatting: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Merge two Excel files and merge/center identical cells in a column.")
    parser.add_argument("file1", help="Path to the first Excel file")
    parser.add_argument("file2", help="Path to the second Excel file")
    parser.add_argument("output", help="Path for the output Excel file")
    
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--col-name", help="Name of the column to merge/center")
    group.add_argument("--col-idx", type=int, help="1-based index of the column to merge/center")

    args = parser.parse_args()

    success = merge_and_format(args.file1, args.file2, args.output, args.col_name, args.col_idx)
    sys.exit(0 if success else 1)
