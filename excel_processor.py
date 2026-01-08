#!/usr/bin/env python3
"""
Excel Processor Module
Handles Excel file manipulation, worksheet updates, refresh operations, and copying.
Designed for the Large Deal Report automation workflow.
"""

import logging
from pathlib import Path
from typing import Optional
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Try to import xlwings for Excel refresh functionality
try:
    import xlwings as xw
    XLWINGS_AVAILABLE = True
except ImportError:
    XLWINGS_AVAILABLE = False
    logging.warning("xlwings not available. Summary refresh will use fallback method.")


class ExcelProcessor:
    """Handles Excel file manipulation and processing operations."""
    
    def __init__(self, config: dict):
        """
        Initialize Excel processor with configuration.
        
        Args:
            config: Dictionary containing configuration:
                - large_deal_report_sheet: Name of the data input sheet (default: 'large deal report')
                - summary_sheet: Name of the summary/pivot sheet (default: 'summary')
                - iphone_compatible_sheet: Name of the iPhone output sheet (default: 'iphone compatible')
        """
        self.config = config
        # Use the actual worksheet names from the user's workbook
        self.large_deal_report_sheet = config.get('large_deal_report_sheet', 'large deal report')
        self.summary_sheet = config.get('summary_sheet', 'summary')
        self.iphone_compatible_sheet = config.get('iphone_compatible_sheet', 'iphone compatible')
        self.logger = logging.getLogger(__name__)
    
    def load_workbook(self, filepath: Path):
        """
        Load an Excel workbook.
        
        Args:
            filepath: Path to the Excel file
            
        Returns:
            openpyxl Workbook object
            
        Raises:
            FileNotFoundError: If file doesn't exist
            Exception: If file cannot be opened
        """
        try:
            self.logger.info(f"Loading workbook: {filepath}")
            return load_workbook(filepath)
        except FileNotFoundError:
            raise FileNotFoundError(f"Workbook not found: {filepath}")
        except Exception as e:
            raise Exception(f"Error loading workbook {filepath}: {str(e)}")
    
    def update_large_deal_report_sheet(self, workbook: openpyxl.Workbook, daily_sheet_path: Path) -> None:
        """
        Replace the 'large deal report' worksheet data with data from the daily email attachment.
        Data is pasted starting from cell A1.
        
        Args:
            workbook: The target workbook to update
            daily_sheet_path: Path to the daily worksheet file (from "large trade td" email)
            
        Raises:
            FileNotFoundError: If daily sheet file doesn't exist
            ValueError: If 'large deal report' worksheet doesn't exist
            Exception: If update fails
        """
        self.logger.info(f"Updating '{self.large_deal_report_sheet}' worksheet...")
        
        try:
            # Load the daily sheet
            if not daily_sheet_path.exists():
                raise FileNotFoundError(f"Daily sheet not found: {daily_sheet_path}")
            
            daily_wb = load_workbook(daily_sheet_path)
            daily_ws = daily_wb.active
            
            # Get the large deal report worksheet
            if self.large_deal_report_sheet in workbook.sheetnames:
                ws = workbook[self.large_deal_report_sheet]
            else:
                # Try case-insensitive match
                ws = None
                for sheet_name in workbook.sheetnames:
                    if sheet_name.lower() == self.large_deal_report_sheet.lower():
                        ws = workbook[sheet_name]
                        self.logger.info(f"Found worksheet with different case: '{sheet_name}'")
                        break
                
                if ws is None:
                    available_sheets = ', '.join(workbook.sheetnames)
                    raise ValueError(
                        f"Worksheet '{self.large_deal_report_sheet}' not found in workbook. "
                        f"Available sheets: {available_sheets}"
                    )
            
            # Clear existing data in the worksheet
            self.logger.info("Clearing existing data...")
            if ws.max_row > 0:
                ws.delete_rows(1, ws.max_row)
            
            # Copy data from daily sheet starting at A1
            self.logger.info("Copying data from daily sheet...")
            for row in daily_ws.iter_rows(values_only=True):
                if any(cell is not None for cell in row):  # Skip completely empty rows
                    ws.append(row)
            
            # Copy formatting (basic cell copying)
            for row_idx, row in enumerate(daily_ws.iter_rows(), start=1):
                if row_idx > ws.max_row:
                    break
                for col_idx, cell in enumerate(row, start=1):
                    if cell.has_style:
                        target_cell = ws.cell(row=row_idx, column=col_idx)
                        target_cell.font = cell.font.copy() if cell.font else None
                        target_cell.fill = cell.fill.copy() if cell.fill else None
                        target_cell.border = cell.border.copy() if cell.border else None
                        target_cell.alignment = cell.alignment.copy() if cell.alignment else None
                        target_cell.number_format = cell.number_format
            
            daily_wb.close()
            self.logger.info(f"'{self.large_deal_report_sheet}' worksheet updated successfully.")
            
        except Exception as e:
            self.logger.error(f"Error updating large deal report sheet: {str(e)}")
            raise
    
    # Alias for backward compatibility
    def update_worksheet_1(self, workbook: openpyxl.Workbook, daily_sheet_path: Path) -> None:
        """Alias for update_large_deal_report_sheet for backward compatibility."""
        return self.update_large_deal_report_sheet(workbook, daily_sheet_path)
    
    def refresh_summary_sheet(self, workbook: openpyxl.Workbook, workbook_path: Optional[Path] = None) -> None:
        """
        Refresh the table/pivot in the 'summary' worksheet.
        This simulates right-clicking and refreshing the table values.
        Uses xlwings (if available) or fallback method.
        
        Args:
            workbook: The workbook containing the summary sheet
            workbook_path: Path to the workbook file (required for xlwings refresh)
            
        Raises:
            ValueError: If summary sheet doesn't exist
            Exception: If refresh fails
        """
        self.logger.info(f"Refreshing '{self.summary_sheet}' worksheet...")
        
        # Check if summary sheet exists (case-insensitive)
        sheet_found = False
        for sheet_name in workbook.sheetnames:
            if sheet_name.lower() == self.summary_sheet.lower():
                sheet_found = True
                break
        
        if not sheet_found:
            available_sheets = ', '.join(workbook.sheetnames)
            raise ValueError(
                f"Worksheet '{self.summary_sheet}' not found in workbook. "
                f"Available sheets: {available_sheets}"
            )
        
        # Try xlwings refresh if available and workbook_path is provided
        if XLWINGS_AVAILABLE and workbook_path:
            try:
                self._refresh_with_xlwings(workbook_path)
                self.logger.info(f"'{self.summary_sheet}' worksheet refreshed using xlwings.")
                return
            except Exception as e:
                self.logger.warning(
                    f"xlwings refresh failed: {str(e)}. Falling back to openpyxl method."
                )
        
        # Fallback: Use openpyxl (preserves structure but doesn't refresh connections)
        self.logger.info(
            "Using openpyxl fallback. Note: Tables, pivot tables, and external data connections "
            "will not be automatically refreshed. You may need to manually refresh in Excel."
        )
        # The worksheet structure is preserved, but data connections won't refresh
        # This is expected behavior when xlwings is not available
    
    # Alias for backward compatibility
    def refresh_worksheet_2(self, workbook: openpyxl.Workbook, workbook_path: Optional[Path] = None) -> None:
        """Alias for refresh_summary_sheet for backward compatibility."""
        return self.refresh_summary_sheet(workbook, workbook_path)
    
    def _refresh_with_xlwings(self, workbook_path: Path) -> None:
        """
        Refresh the summary worksheet using xlwings (requires Excel installed).
        This simulates the right-click -> refresh action on tables/pivot tables.
        Works on Windows (primarily designed for Windows 11 with Excel).
        
        Args:
            workbook_path: Path to the workbook file
            
        Raises:
            Exception: If xlwings refresh fails
        """
        app = None
        wb = None
        
        try:
            self.logger.info("Attempting to refresh using xlwings (Excel automation)...")
            
            # Open Excel application (headless on Windows)
            app = xw.App(visible=False)
            app.display_alerts = False
            
            # Open the workbook
            wb = app.books.open(str(workbook_path))
            
            # Find the summary sheet (case-insensitive)
            summary_ws = None
            for sheet in wb.sheets:
                if sheet.name.lower() == self.summary_sheet.lower():
                    summary_ws = sheet
                    break
            
            if summary_ws is None:
                raise ValueError(f"Worksheet '{self.summary_sheet}' not found")
            
            # Try to refresh all data connections/queries in the workbook
            # This is equivalent to right-click -> refresh on tables
            try:
                # RefreshAll refreshes all data connections, pivot tables, and queries
                wb.api.RefreshAll()
                self.logger.info("Refreshed all data connections in workbook.")
            except Exception as e:
                self.logger.warning(f"Could not use RefreshAll: {str(e)}")
                # Try to refresh specific tables (ListObjects) on the summary sheet
                try:
                    if summary_ws.api.ListObjects.Count > 0:
                        for i in range(1, summary_ws.api.ListObjects.Count + 1):
                            list_obj = summary_ws.api.ListObjects(i)
                            self.logger.info(f"Refreshing table: {list_obj.Name}")
                            if hasattr(list_obj, 'QueryTable') and list_obj.QueryTable is not None:
                                list_obj.QueryTable.Refresh(BackgroundQuery=False)
                            else:
                                # For regular tables, try to refresh if linked to a source
                                try:
                                    list_obj.Refresh()
                                except:
                                    pass
                        self.logger.info("Refreshed table(s) on summary sheet.")
                except Exception as e2:
                    self.logger.warning(f"Could not refresh tables: {str(e2)}")
                
                # Try to refresh pivot tables
                try:
                    if summary_ws.api.PivotTables.Count > 0:
                        for i in range(1, summary_ws.api.PivotTables.Count + 1):
                            pivot = summary_ws.api.PivotTables(i)
                            self.logger.info(f"Refreshing pivot table: {pivot.Name}")
                            pivot.RefreshTable()
                        self.logger.info("Refreshed pivot table(s) on summary sheet.")
                except Exception as e3:
                    self.logger.warning(f"Could not refresh pivot tables: {str(e3)}")
            
            # Save the workbook
            wb.save()
            self.logger.info("Workbook saved after refresh.")
            
        except Exception as e:
            self.logger.error(f"xlwings refresh error: {str(e)}")
            raise
        finally:
            # Clean up
            if wb:
                try:
                    wb.close()
                except:
                    pass
            if app:
                try:
                    app.quit()
                except:
                    pass
    
    def copy_summary_to_iphone_compatible(self, workbook: openpyxl.Workbook) -> None:
        """
        Copy the table from 'summary' sheet (starting at B6) to 'iphone compatible' sheet (starting at A1).
        The table starts at B6 with headers: Fund name, Register name, etc.
        Skips the macro buttons in B2-B4.
        
        Args:
            workbook: The workbook containing the worksheets
            
        Raises:
            ValueError: If summary sheet doesn't exist
            Exception: If copy operation fails
        """
        self.logger.info(f"Copying table from '{self.summary_sheet}' to '{self.iphone_compatible_sheet}'...")
        
        # Find the summary sheet (case-insensitive)
        summary_ws = None
        for sheet_name in workbook.sheetnames:
            if sheet_name.lower() == self.summary_sheet.lower():
                summary_ws = workbook[sheet_name]
                break
        
        if summary_ws is None:
            available_sheets = ', '.join(workbook.sheetnames)
            raise ValueError(
                f"Worksheet '{self.summary_sheet}' not found in workbook. "
                f"Available sheets: {available_sheets}"
            )
        
        try:
            # Find or create the iphone compatible sheet
            iphone_ws = None
            for sheet_name in workbook.sheetnames:
                if sheet_name.lower() == self.iphone_compatible_sheet.lower():
                    iphone_ws = workbook[sheet_name]
                    # Clear existing content
                    self.logger.info(f"Clearing existing '{sheet_name}' sheet...")
                    if iphone_ws.max_row > 0:
                        iphone_ws.delete_rows(1, iphone_ws.max_row)
                    break
            
            if iphone_ws is None:
                # Create new iphone compatible sheet
                self.logger.info(f"Creating new '{self.iphone_compatible_sheet}' sheet...")
                iphone_ws = workbook.create_sheet(self.iphone_compatible_sheet)
            
            # Table starts at B6 in summary sheet
            # Copy from B6 onwards to A1 onwards in iphone compatible
            start_row = 6  # Table starts at row 6
            start_col = 2  # Column B = 2
            
            self.logger.info(f"Copying table from B{start_row} onwards...")
            
            # Find the last row and column with data
            max_row = summary_ws.max_row
            max_col = summary_ws.max_column
            
            self.logger.info(f"Table range: B{start_row} to {get_column_letter(max_col)}{max_row}")
            
            # Copy the table data
            dest_row = 1
            for src_row in range(start_row, max_row + 1):
                dest_col = 1
                row_has_data = False
                
                for src_col in range(start_col, max_col + 1):
                    source_cell = summary_ws.cell(row=src_row, column=src_col)
                    target_cell = iphone_ws.cell(row=dest_row, column=dest_col)
                    
                    # Copy the value
                    target_cell.value = source_cell.value
                    
                    if source_cell.value is not None:
                        row_has_data = True
                    
                    # Copy formatting
                    if source_cell.has_style:
                        target_cell.font = source_cell.font.copy() if source_cell.font else None
                        target_cell.fill = source_cell.fill.copy() if source_cell.fill else None
                        target_cell.border = source_cell.border.copy() if source_cell.border else None
                        target_cell.alignment = source_cell.alignment.copy() if source_cell.alignment else None
                        target_cell.number_format = source_cell.number_format
                    
                    dest_col += 1
                
                # Only increment destination row if source row had data
                if row_has_data:
                    dest_row += 1
            
            # Copy column widths (shifted from B->A, C->B, etc.)
            for src_col in range(start_col, max_col + 1):
                src_letter = get_column_letter(src_col)
                dest_letter = get_column_letter(src_col - start_col + 1)
                if src_letter in summary_ws.column_dimensions:
                    iphone_ws.column_dimensions[dest_letter].width = summary_ws.column_dimensions[src_letter].width
            
            self.logger.info(f"'{self.iphone_compatible_sheet}' sheet created successfully.")
            self.logger.info(f"Copied {dest_row - 1} rows of data.")
            
        except Exception as e:
            self.logger.error(f"Error copying summary to iphone compatible: {str(e)}")
            raise
    
    # Alias for backward compatibility
    def copy_worksheet_2_to_3(self, workbook: openpyxl.Workbook) -> None:
        """Alias for copy_summary_to_iphone_compatible for backward compatibility."""
        return self.copy_summary_to_iphone_compatible(workbook)
    
    def save_workbook(self, workbook: openpyxl.Workbook, filepath: Path) -> None:
        """
        Save the workbook to a file.
        
        Args:
            workbook: The workbook to save
            filepath: Path where to save the workbook
            
        Raises:
            Exception: If save fails
        """
        try:
            self.logger.info(f"Saving workbook to: {filepath}")
            workbook.save(filepath)
            self.logger.info("Workbook saved successfully.")
        except Exception as e:
            self.logger.error(f"Error saving workbook: {str(e)}")
            raise
