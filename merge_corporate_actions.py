#!/usr/bin/env python3
"""
Corporate Actions Data Merge Utility

Merges data from two Excel spreadsheets, normalizing column names,
extracting specific columns, sorting, and applying hierarchical merge & center formatting.
"""

import logging
import argparse
from pathlib import Path
from typing import List, Dict, Optional, Tuple
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Target columns in order
TARGET_COLUMNS = [
    'ISSUE NAME',
    'SEDOL', 
    'ISIN',
    'ISSUE COUNTRY NAME',
    'STRATEGY',
    'SUB STRATEGY',
    'FUND TYPE',
    'FUND NAME'
]

# Columns that should NOT establish a new scope for subsequent columns
# ISIN is unique, so we don't want it to prevent merging of Country, Strategy, etc.
NON_SCOPING_COLUMNS = ['ISIN']

def normalize_and_extract(df: pd.DataFrame, source_name: str) -> pd.DataFrame:
    """
    Normalizes DataFrame columns to uppercase and extracts target columns.
    """
    # Create a mapping of {upper_case_name: original_name}
    col_map = {c.strip().upper(): c for c in df.columns}
    
    selected_data = {}
    missing_cols = []
    
    for target in TARGET_COLUMNS:
        if target in col_map:
            selected_data[target] = df[col_map[target]]
        else:
            missing_cols.append(target)
    
    if missing_cols:
        logger.warning(f"Missing columns in {source_name}: {missing_cols}")
        for col in missing_cols:
            selected_data[col] = pd.Series([None] * len(df))
            
    return pd.DataFrame(selected_data)

def get_merge_ranges(series: pd.Series, parent_ranges: Optional[List[Tuple[int, int]]] = None) -> List[Tuple[int, int]]:
    """
    Identifies ranges of consecutive identical values.
    If parent_ranges is provided, only merges within the boundaries of parent ranges.
    """
    ranges = []
    
    if parent_ranges is None:
        parent_ranges = [(0, len(series) - 1)]
    
    for p_start, p_end in parent_ranges:
        current_start = p_start
        # Handle empty series case
        if p_start >= len(series):
            continue
            
        current_val = series.iloc[p_start]
        
        for i in range(p_start + 1, p_end + 1):
            val = series.iloc[i]
            
            is_different = False
            if pd.isna(current_val) and pd.isna(val):
                is_different = False
            elif pd.isna(current_val) or pd.isna(val):
                is_different = True
            elif current_val != val:
                is_different = True
                
            if is_different:
                if i - 1 > current_start:
                    ranges.append((current_start, i - 1))
                current_start = i
                current_val = val
        
        if p_end > current_start:
            ranges.append((current_start, p_end))
            
    return ranges

def apply_hierarchical_merges(ws, df: pd.DataFrame):
    """
    Applies merge and center formatting based on hierarchical grouping.
    """
    ROW_OFFSET = 2
    parent_ranges = None
    
    for col_idx, col_name in enumerate(TARGET_COLUMNS):
        series = df[col_name]
        
        # Get ranges to merge for this column using the current parent scope
        current_ranges = get_merge_ranges(series, parent_ranges)
        
        # Apply merges to Excel
        excel_col_idx = col_idx + 1
        col_letter = get_column_letter(excel_col_idx)
        
        for start_idx, end_idx in current_ranges:
            start_row = start_idx + ROW_OFFSET
            end_row = end_idx + ROW_OFFSET
            
            # Only merge if it spans multiple rows
            if end_row > start_row:
                merge_ref = f"{col_letter}{start_row}:{col_letter}{end_row}"
                try:
                    ws.merge_cells(merge_ref)
                    cell = ws.cell(row=start_row, column=excel_col_idx)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                except Exception as e:
                    logger.error(f"Failed to merge {merge_ref}: {e}")
        
        # Update parent ranges for the next column, UNLESS this column is non-scoping
        if col_name not in NON_SCOPING_COLUMNS:
            parent_ranges = current_ranges

def merge_spreadsheets(file1: str, file2: str, output_file: str):
    logger.info(f"Reading {file1}...")
    df1 = pd.read_excel(file1)
    
    logger.info(f"Reading {file2}...")
    df2 = pd.read_excel(file2)
    
    logger.info("Normalizing columns...")
    df1_norm = normalize_and_extract(df1, "File 1")
    df2_norm = normalize_and_extract(df2, "File 2")
    
    logger.info("Merging data...")
    merged_df = pd.concat([df1_norm, df2_norm], ignore_index=True)
    
    logger.info("Sorting data...")
    # Sort by hierarchy
    merged_df = merged_df.sort_values(by=TARGET_COLUMNS, na_position='last')
    merged_df = merged_df.reset_index(drop=True)
    
    logger.info("Writing to Excel...")
    merged_df.to_excel(output_file, index=False)
    
    logger.info("Applying formatting...")
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    
    apply_hierarchical_merges(ws, merged_df)
    
    wb.save(output_file)
    logger.info(f"Done! Saved to {output_file}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Merge corporate actions spreadsheets.")
    parser.add_argument("file1", help="First Excel file path")
    parser.add_argument("file2", help="Second Excel file path")
    parser.add_argument("output", help="Output Excel file path")
    
    args = parser.parse_args()
    merge_spreadsheets(args.file1, args.file2, args.output)
